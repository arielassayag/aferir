"""Receita de referência da União (LC 214, art. 350, I): fetch + parse.

Rota primária: XLSX "Arrecadação das receitas federais 1994 a 2025"
(RFB/gov.br, R$ milhões correntes, uma aba por ano, linhas por tributo).
Fallback e validação cruzada: IPEADATA OData4 (séries mensais SRF).

Notas de fronteira (C3):
 - PIS × Pasep: a série aberta publica só a linha conjunta "CONTRIBUIÇÃO PARA
   O PIS/PASEP". Não há perda: o art. 350, I, 'a' referencia a contribuição
   por inteiro (CF, art. 239). Documentado aqui e em r_uniao.csv.
 - IOF-Seguros: nenhuma fonte aberta máquina-legível decompõe o IOF por
   modalidade (verificado: XLSX RFB, Portal da Transparência, Análise da
   Arrecadação dez/2025). O rateio fica em aferir.inputs.uniao (razão de
   quadro publicado — NT SERT jul/2024, p. 3).

Idempotência: cache em data/raw/{rfb,ipeadata}/ com _meta.json
(url, sha256, collected_at). datetime SÓ aqui (fetcher), nunca no cálculo.
"""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
import requests

from aferir import config
from aferir.provenance import sha256_file

RAW_RFB = config.RAW / "rfb"
RAW_IPEADATA = config.RAW / "ipeadata"
RAW_STN = config.RAW / "stn"

_UA = {"User-Agent": "Mozilla/5.0 (aferir; rotina publica de dados abertos)"}

# Coluna do TOTAL anual no XLSX (col. N; JAN..DEZ em B..M) — verificado 2012-2025.
_COL_TOTAL = 13
_N_MESES = 12


def _download(url: str, destino: Path, *, force: bool = False) -> Path:
    """Baixa `url` para `destino` (idempotente) e grava `<destino>._meta.json`."""
    destino.parent.mkdir(parents=True, exist_ok=True)
    meta_path = destino.with_name(destino.name + "._meta.json")
    if destino.exists() and meta_path.exists() and not force:
        return destino
    resp = requests.get(url, headers=_UA, timeout=180)
    resp.raise_for_status()
    destino.write_bytes(resp.content)
    meta = {
        "url": url,
        "sha256": sha256_file(destino),
        "bytes": destino.stat().st_size,
        "collected_at": datetime.now(timezone.utc).isoformat(),
    }
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=1) + "\n",
                         encoding="utf-8")
    return destino


def fetch_rfb_xlsx(*, force: bool = False) -> Path:
    """XLSX da série histórica de arrecadação federal (rota primária)."""
    nome = config.RFB_XLSX_URL.rsplit("/", 1)[-1]
    return _download(config.RFB_XLSX_URL, RAW_RFB / nome, force=force)


def fetch_ipeadata(tributo: str, *, force: bool = False) -> Path:
    """Série mensal IPEADATA OData4 do tributo (fallback/validação cruzada)."""
    sercodigo = config.IPEADATA_SERCODIGOS[tributo]
    url = config.IPEADATA_ODATA_URL.format(sercodigo=sercodigo)
    return _download(url, RAW_IPEADATA / f"{sercodigo}.json", force=force)


def fetch_all(*, force: bool = False) -> list[Path]:
    """Baixa XLSX RFB + XLSX RTN/STN + 4 séries IPEADATA (idempotente)."""
    paths = [fetch_rfb_xlsx(force=force), fetch_rtn_xlsx(force=force)]
    for tributo in config.IPEADATA_SERCODIGOS:
        paths.append(fetch_ipeadata(tributo, force=force))
    return paths


# ------------------------------------------------------------------ parse
def parse_rfb_receitas(anos: list[int] | None = None,
                       path: Path | None = None) -> pd.DataFrame:
    """Extrai [ano × tributo] (R$ mi correntes) do XLSX RFB.

    Usa a coluna TOTAL de cada aba-ano, validando que o cabeçalho tem os 12
    meses + TOTAL e que TOTAL == soma dos meses (1e-6 relativo).
    """
    import openpyxl

    if path is None:
        path = fetch_rfb_xlsx()
    if anos is None:
        anos = sorted(set(config.ANCORA_UNIAO) | set(config.JANELA_RECEITA))

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    registros = []
    rotulos = {v: k for k, v in config.RFB_XLSX_ROTULOS.items()}
    for ano in anos:
        ws = wb[str(ano)]
        achados: dict[str, float] = {}
        for row in ws.iter_rows(values_only=True):
            rotulo = str(row[0]).strip() if row[0] is not None else ""
            if rotulo == "RECEITAS":  # linha de cabeçalho JAN..DEZ TOTAL
                meses = [c for c in row[1:] if c not in (None, "")]
                if len(meses) != _N_MESES + 1 or str(meses[-1]).strip() != "TOTAL":
                    raise ValueError(f"{ano}: cabeçalho inesperado: {meses}")
            if rotulo in rotulos:
                total = float(row[_COL_TOTAL])
                soma_meses = sum(float(c) for c in row[1:1 + _N_MESES])
                if abs(total - soma_meses) > 1e-6 * max(abs(total), 1.0):
                    raise ValueError(
                        f"{ano}/{rotulo}: TOTAL {total} != soma meses {soma_meses}")
                achados[rotulos[rotulo]] = total
        faltando = set(config.RFB_XLSX_ROTULOS) - set(achados)
        if faltando:
            raise ValueError(f"{ano}: linhas ausentes no XLSX: {faltando}")
        registros.append({"ano": ano, **achados})
    wb.close()
    df = pd.DataFrame(registros).sort_values("ano").reset_index(drop=True)
    return df[["ano", *config.RFB_XLSX_ROTULOS]]


def fetch_rtn_xlsx(*, force: bool = False) -> Path:
    """XLSX 'Resultado do Tesouro Nacional — Série Histórica' (STN/CKAN).

    Rota aberta da convenção LÍQUIDA de restituições por tributo (Tema 69):
    Tabela 2.2 (anual), conceito caixa — ingresso efetivo na Conta Única.
    """
    nome = config.RTN_XLSX_URL.rsplit("/", 1)[-1]
    return _download(config.RTN_XLSX_URL, RAW_STN / nome, force=force)


def parse_rtn_receitas(anos: list[int] | None = None,
                       path: Path | None = None) -> pd.DataFrame:
    """Extrai [ano × tributo] (R$ mi correntes, LÍQUIDA-RTN) da Tabela 2.2.

    Localiza a linha de cabeçalho ('Discriminação' + anos) e as linhas dos
    tributos pelos prefixos de config.RTN_ROTULOS (rótulos '1.1.xx  Nome').
    Valida presença de todos os anos e tributos pedidos.
    """
    import openpyxl

    if path is None:
        path = fetch_rtn_xlsx()
    if anos is None:
        anos = list(config.ANCORA_UNIAO)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[config.RTN_ABA_ANUAL]
    linhas = list(ws.iter_rows(values_only=True))
    wb.close()

    col_ano: dict[int, int] = {}
    for row in linhas:
        r0 = str(row[0]).strip() if row[0] is not None else ""
        if r0.startswith("Discriminaç"):
            col_ano = {int(str(v).strip()): j
                       for j, v in enumerate(row[1:], start=1)
                       if v is not None and str(v).strip().isdigit()}
            break
    faltam = set(anos) - set(col_ano)
    if not col_ano or faltam:
        raise ValueError(f"RTN 2.2: cabeçalho/anos ausentes: {sorted(faltam)}")

    achados: dict[str, dict[int, float]] = {}
    for row in linhas:
        r0 = str(row[0]).strip() if row[0] is not None else ""
        for tributo, prefixo in config.RTN_ROTULOS.items():
            if r0.startswith(prefixo) and tributo not in achados:
                achados[tributo] = {a: float(row[j]) for a, j in col_ano.items()
                                    if a in anos}
    faltando = set(config.RTN_ROTULOS) - set(achados)
    if faltando:
        raise ValueError(f"RTN 2.2: linhas ausentes: {faltando}")

    registros = [{"ano": a, **{t: achados[t][a] for t in config.RTN_ROTULOS}}
                 for a in sorted(anos)]
    return pd.DataFrame(registros)[["ano", *config.RTN_ROTULOS]]


def ipeadata_anual(tributo: str, anos: list[int]) -> pd.Series:
    """Anualiza (soma dos meses) a série IPEADATA do tributo, em R$ mi."""
    path = fetch_ipeadata(tributo)
    dados = json.loads(path.read_text(encoding="utf-8"))["value"]
    por_ano: dict[int, float] = {}
    n_meses: dict[int, int] = {}
    for v in dados:
        ano = int(v["VALDATA"][:4])
        if ano in anos and v["VALVALOR"] is not None:
            por_ano[ano] = por_ano.get(ano, 0.0) + float(v["VALVALOR"])
            n_meses[ano] = n_meses.get(ano, 0) + 1
    incompletos = {a: n for a, n in n_meses.items() if n != _N_MESES}
    if incompletos:
        raise ValueError(f"IPEADATA {tributo}: anos incompletos {incompletos}")
    faltantes = set(anos) - set(por_ano)
    if faltantes:
        raise ValueError(f"IPEADATA {tributo}: anos ausentes {sorted(faltantes)}")
    return pd.Series(por_ano, name=tributo).sort_index()


def crosscheck_ipeadata(anos: list[int] | None = None,
                        tol_rel: float = 1e-3) -> pd.DataFrame:
    """Confronta XLSX RFB × IPEADATA ano a ano; falha se divergir > tol_rel.

    Retorna o quadro de diferenças relativas (auditoria da triangulação).
    """
    if anos is None:
        anos = sorted(set(config.ANCORA_UNIAO) | set(config.JANELA_RECEITA))
    xlsx = parse_rfb_receitas(anos).set_index("ano")
    linhas = []
    for tributo in config.IPEADATA_SERCODIGOS:
        ipea = ipeadata_anual(tributo, anos)
        for ano in anos:
            a, b = float(xlsx.loc[ano, tributo]), float(ipea[ano])
            diff = abs(a - b) / max(abs(a), 1.0)
            linhas.append({"ano": ano, "tributo": tributo, "xlsx_rs_mi": a,
                           "ipeadata_rs_mi": b, "diff_rel": diff})
    quadro = pd.DataFrame(linhas)
    pior = quadro["diff_rel"].max()
    if pior > tol_rel:
        ruins = quadro[quadro["diff_rel"] > tol_rel]
        raise AssertionError(f"XLSX × IPEADATA divergem além de {tol_rel}:\n{ruins}")
    return quadro
